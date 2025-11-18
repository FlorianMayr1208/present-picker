from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from werkzeug.utils import secure_filename
from app.config import Config
from app.models import db
from app.models.destination import Destination
from app.models.activity import Activity
import os
import csv
import io
from openpyxl import Workbook, load_workbook
from datetime import datetime


def create_app(config_class=Config):
    """Application Factory Pattern"""
    app = Flask(__name__)
    app.config.from_object(config_class)

    # Initialisiere Extensions
    db.init_app(app)

    # Erstelle Datenbank-Tabellen nur in Development Mode
    # Für Production sollte ein separates Migrations-Skript verwendet werden
    with app.app_context():
        try:
            # Prüfe ob Tabellen existieren, ohne sie zu erstellen
            db.engine.connect()
        except Exception as e:
            # Falls Verbindung fehlschlägt, logge den Fehler aber fahre fort
            app.logger.warning(f"Database connection check failed: {e}")

    # Registriere Routen
    register_routes(app)

    return app


def register_routes(app):
    """Registriere alle Routes"""

    @app.route('/')
    def index():
        """Startseite mit allen Destinationen"""
        destinations = Destination.query.all()
        return render_template('index.html', destinations=destinations)

    @app.route('/destination/<int:id>')
    def show_destination(id):
        """Detailansicht einer Destination mit Slider"""
        slider_value = request.args.get('slider', 0, type=int)
        destination = Destination.query.get_or_404(id)

        # Filtere Aktivitäten basierend auf Slider-Level Range
        activities = Activity.query.filter(
            Activity.destination_id == id,
            Activity.slider_level_min <= slider_value,
            Activity.slider_level_max >= slider_value
        ).order_by(Activity.slider_level_min).all()

        return render_template(
            'destination.html',
            destination=destination,
            activities=activities,
            slider_value=slider_value,
            slider_max=app.config['SLIDER_MAX']
        )

    @app.route('/api/destination/<int:id>/activities')
    def api_get_activities(id):
        """API-Endpunkt: Hole Aktivitäten basierend auf Slider-Level"""
        from flask import jsonify

        slider_value = request.args.get('slider', 0, type=int)
        destination = Destination.query.get_or_404(id)

        # Filtere Aktivitäten basierend auf Slider-Level Range
        activities = Activity.query.filter(
            Activity.destination_id == id,
            Activity.slider_level_min <= slider_value,
            Activity.slider_level_max >= slider_value
        ).order_by(Activity.slider_level_min).all()

        # Konvertiere zu JSON
        activities_data = [{
            'id': activity.id,
            'title': activity.title,
            'description': activity.description,
            'slider_level_min': activity.slider_level_min,
            'slider_level_max': activity.slider_level_max,
            'image_filename': activity.image_filename
        } for activity in activities]

        return jsonify({
            'activities': activities_data,
            'count': len(activities_data)
        })

    @app.route('/admin')
    def admin():
        """Admin-Übersicht"""
        destinations = Destination.query.all()
        return render_template('admin.html', destinations=destinations)

    # ========== DESTINATION ADMIN ROUTES ==========

    @app.route('/admin/destination/new', methods=['GET', 'POST'])
    def admin_destination_new():
        """Neue Destination erstellen"""
        if request.method == 'POST':
            name = request.form.get('name')
            description_short = request.form.get('description_short')
            image_file = request.files.get('image_cover')

            # Validierung
            if not name:
                flash('Name ist erforderlich!', 'danger')
                return render_template('admin_destination_form.html', destination=None)

            # Bild-Upload verarbeiten
            image_filename = None
            if image_file and image_file.filename:
                image_filename = save_uploaded_file(image_file, 'general')

            # Destination erstellen
            destination = Destination(
                name=name,
                description_short=description_short,
                image_cover=image_filename
            )
            db.session.add(destination)
            db.session.commit()

            flash(f'Destination "{name}" wurde erfolgreich erstellt!', 'success')
            return redirect(url_for('admin'))

        return render_template('admin_destination_form.html', destination=None)

    @app.route('/admin/destination/<int:id>/edit', methods=['GET', 'POST'])
    def admin_destination_edit(id):
        """Destination bearbeiten"""
        destination = Destination.query.get_or_404(id)

        if request.method == 'POST':
            destination.name = request.form.get('name')
            destination.description_short = request.form.get('description_short')
            image_file = request.files.get('image_cover')

            # Validierung
            if not destination.name:
                flash('Name ist erforderlich!', 'danger')
                return render_template('admin_destination_form.html', destination=destination)

            # Neues Bild hochladen (optional)
            if image_file and image_file.filename:
                destination.image_cover = save_uploaded_file(image_file, 'general')

            db.session.commit()
            flash(f'Destination "{destination.name}" wurde aktualisiert!', 'success')
            return redirect(url_for('admin'))

        return render_template('admin_destination_form.html', destination=destination)

    @app.route('/admin/destination/<int:id>/delete', methods=['POST'])
    def admin_destination_delete(id):
        """Destination löschen"""
        destination = Destination.query.get_or_404(id)
        name = destination.name
        db.session.delete(destination)
        db.session.commit()
        flash(f'Destination "{name}" wurde gelöscht!', 'warning')
        return redirect(url_for('admin'))

    # ========== ACTIVITY ADMIN ROUTES ==========

    @app.route('/admin/destination/<int:destination_id>/activities')
    def admin_activities(destination_id):
        """Aktivitäten einer Destination verwalten"""
        destination = Destination.query.get_or_404(destination_id)
        activities = Activity.query.filter_by(destination_id=destination_id).order_by(Activity.slider_level_min).all()
        return render_template('admin_activities.html', destination=destination, activities=activities)

    @app.route('/admin/destination/<int:destination_id>/activity/new', methods=['GET', 'POST'])
    def admin_activity_new(destination_id):
        """Neue Aktivität erstellen"""
        destination = Destination.query.get_or_404(destination_id)

        if request.method == 'POST':
            title = request.form.get('title')
            description = request.form.get('description')
            slider_level_min = request.form.get('slider_level_min', type=int)
            slider_level_max = request.form.get('slider_level_max', type=int)
            image_file = request.files.get('image_filename')

            # Validierung
            if not title:
                flash('Titel ist erforderlich!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=None,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_min is None or slider_level_min < 0 or slider_level_min > app.config['SLIDER_MAX']:
                flash(f'Minimales Slider-Level muss zwischen 0 und {app.config["SLIDER_MAX"]} liegen!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=None,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_max is None or slider_level_max < 0 or slider_level_max > app.config['SLIDER_MAX']:
                flash(f'Maximales Slider-Level muss zwischen 0 und {app.config["SLIDER_MAX"]} liegen!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=None,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_min > slider_level_max:
                flash('Minimales Level darf nicht größer als maximales Level sein!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=None,
                                     slider_max=app.config['SLIDER_MAX'])

            # Bild-Upload verarbeiten
            image_filename = None
            if image_file and image_file.filename:
                # Erstelle Ordner für diese Destination falls nicht vorhanden
                dest_folder = sanitize_folder_name(destination.name)
                image_filename = save_uploaded_file(image_file, dest_folder)

            # Aktivität erstellen
            activity = Activity(
                destination_id=destination_id,
                title=title,
                description=description,
                slider_level_min=slider_level_min,
                slider_level_max=slider_level_max,
                image_filename=image_filename
            )
            db.session.add(activity)
            db.session.commit()

            flash(f'Aktivität "{title}" wurde erstellt!', 'success')
            return redirect(url_for('admin_activities', destination_id=destination_id))

        return render_template('admin_activity_form.html',
                             destination=destination,
                             activity=None,
                             slider_max=app.config['SLIDER_MAX'])

    @app.route('/admin/activity/<int:id>/edit', methods=['GET', 'POST'])
    def admin_activity_edit(id):
        """Aktivität bearbeiten"""
        activity = Activity.query.get_or_404(id)
        destination = activity.destination

        if request.method == 'POST':
            activity.title = request.form.get('title')
            activity.description = request.form.get('description')
            slider_level_min = request.form.get('slider_level_min', type=int)
            slider_level_max = request.form.get('slider_level_max', type=int)
            image_file = request.files.get('image_filename')

            # Validierung
            if not activity.title:
                flash('Titel ist erforderlich!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=activity,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_min is None or slider_level_min < 0 or slider_level_min > app.config['SLIDER_MAX']:
                flash(f'Minimales Slider-Level muss zwischen 0 und {app.config["SLIDER_MAX"]} liegen!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=activity,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_max is None or slider_level_max < 0 or slider_level_max > app.config['SLIDER_MAX']:
                flash(f'Maximales Slider-Level muss zwischen 0 und {app.config["SLIDER_MAX"]} liegen!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=activity,
                                     slider_max=app.config['SLIDER_MAX'])

            if slider_level_min > slider_level_max:
                flash('Minimales Level darf nicht größer als maximales Level sein!', 'danger')
                return render_template('admin_activity_form.html',
                                     destination=destination,
                                     activity=activity,
                                     slider_max=app.config['SLIDER_MAX'])

            activity.slider_level_min = slider_level_min
            activity.slider_level_max = slider_level_max

            # Neues Bild hochladen (optional)
            if image_file and image_file.filename:
                dest_folder = sanitize_folder_name(destination.name)
                activity.image_filename = save_uploaded_file(image_file, dest_folder)

            db.session.commit()
            flash(f'Aktivität "{activity.title}" wurde aktualisiert!', 'success')
            return redirect(url_for('admin_activities', destination_id=destination.id))

        return render_template('admin_activity_form.html',
                             destination=destination,
                             activity=activity,
                             slider_max=app.config['SLIDER_MAX'])

    @app.route('/admin/activity/<int:id>/delete', methods=['POST'])
    def admin_activity_delete(id):
        """Aktivität löschen"""
        activity = Activity.query.get_or_404(id)
        destination_id = activity.destination_id
        title = activity.title
        db.session.delete(activity)
        db.session.commit()
        flash(f'Aktivität "{title}" wurde gelöscht!', 'warning')
        return redirect(url_for('admin_activities', destination_id=destination_id))

    # ========== IMPORT/EXPORT ROUTES ==========

    @app.route('/admin/export/xlsx')
    def export_xlsx():
        """Exportiere alle Daten als Excel-Datei"""
        wb = Workbook()

        # Sheet 1: Destinationen
        ws_dest = wb.active
        ws_dest.title = "Destinationen"
        ws_dest.append(['ID', 'Name', 'Kurzbeschreibung', 'Titelbild'])

        destinations = Destination.query.all()
        for dest in destinations:
            ws_dest.append([
                dest.id,
                dest.name,
                dest.description_short or '',
                dest.image_cover or ''
            ])

        # Sheet 2: Aktivitäten
        ws_act = wb.create_sheet("Aktivitäten")
        ws_act.append(['ID', 'Destination ID', 'Destination Name', 'Titel', 'Beschreibung', 'Slider Level Min', 'Slider Level Max', 'Bild'])

        activities = Activity.query.all()
        for act in activities:
            ws_act.append([
                act.id,
                act.destination_id,
                act.destination.name,
                act.title,
                act.description or '',
                act.slider_level_min,
                act.slider_level_max,
                act.image_filename or ''
            ])

        # Speichere in BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'reiseplanung_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    @app.route('/admin/export/csv')
    def export_csv():
        """Exportiere alle Daten als CSV-Datei (ZIP mit 2 Dateien)"""
        import zipfile

        # Erstelle ZIP im Speicher
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # CSV 1: Destinationen
            dest_output = io.StringIO()
            dest_writer = csv.writer(dest_output)
            dest_writer.writerow(['ID', 'Name', 'Kurzbeschreibung', 'Titelbild'])

            destinations = Destination.query.all()
            for dest in destinations:
                dest_writer.writerow([
                    dest.id,
                    dest.name,
                    dest.description_short or '',
                    dest.image_cover or ''
                ])

            zip_file.writestr('destinationen.csv', dest_output.getvalue())

            # CSV 2: Aktivitäten
            act_output = io.StringIO()
            act_writer = csv.writer(act_output)
            act_writer.writerow(['ID', 'Destination ID', 'Destination Name', 'Titel', 'Beschreibung', 'Slider Level Min', 'Slider Level Max', 'Bild'])

            activities = Activity.query.all()
            for act in activities:
                act_writer.writerow([
                    act.id,
                    act.destination_id,
                    act.destination.name,
                    act.title,
                    act.description or '',
                    act.slider_level_min,
                    act.slider_level_max,
                    act.image_filename or ''
                ])

            zip_file.writestr('aktivitaeten.csv', act_output.getvalue())

        zip_buffer.seek(0)
        filename = f'reiseplanung_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )

    @app.route('/admin/import', methods=['GET', 'POST'])
    def import_data():
        """Importiere Daten aus Excel oder CSV"""
        if request.method == 'POST':
            file = request.files.get('import_file')

            if not file or file.filename == '':
                flash('Keine Datei ausgewählt!', 'danger')
                return redirect(url_for('import_data'))

            filename = secure_filename(file.filename)

            try:
                if filename.endswith('.xlsx'):
                    # Excel Import
                    wb = load_workbook(file)

                    # Importiere Destinationen
                    if 'Destinationen' in wb.sheetnames:
                        ws_dest = wb['Destinationen']
                        rows = list(ws_dest.iter_rows(min_row=2, values_only=True))

                        for row in rows:
                            if row[1]:  # Wenn Name vorhanden
                                dest = Destination(
                                    name=row[1],
                                    description_short=row[2] if row[2] else None,
                                    image_cover=row[3] if row[3] else None
                                )
                                db.session.add(dest)

                        db.session.commit()
                        flash(f'{len(rows)} Destination(en) importiert!', 'success')

                    # Importiere Aktivitäten
                    if 'Aktivitäten' in wb.sheetnames:
                        ws_act = wb['Aktivitäten']
                        rows = list(ws_act.iter_rows(min_row=2, values_only=True))

                        for row in rows:
                            if row[1] and row[3]:  # Wenn Destination ID und Titel vorhanden
                                activity = Activity(
                                    destination_id=row[1],
                                    title=row[3],
                                    description=row[4] if row[4] else None,
                                    slider_level_min=row[5] if row[5] else 0,
                                    slider_level_max=row[6] if row[6] else 5,
                                    image_filename=row[7] if row[7] else None
                                )
                                db.session.add(activity)

                        db.session.commit()
                        flash(f'{len(rows)} Aktivität(en) importiert!', 'success')

                elif filename.endswith('.csv'):
                    # CSV Import
                    content = file.read().decode('utf-8')
                    csv_reader = csv.DictReader(io.StringIO(content))

                    # Erkenne, ob es Destinationen oder Aktivitäten sind
                    rows = list(csv_reader)
                    if rows and 'Name' in rows[0]:
                        # Destinationen
                        for row in rows:
                            dest = Destination(
                                name=row['Name'],
                                description_short=row.get('Kurzbeschreibung', ''),
                                image_cover=row.get('Titelbild', '')
                            )
                            db.session.add(dest)
                        db.session.commit()
                        flash(f'{len(rows)} Destination(en) aus CSV importiert!', 'success')

                    elif rows and 'Titel' in rows[0]:
                        # Aktivitäten
                        for row in rows:
                            activity = Activity(
                                destination_id=int(row['Destination ID']),
                                title=row['Titel'],
                                description=row.get('Beschreibung', ''),
                                slider_level_min=int(row.get('Slider Level Min', 0)),
                                slider_level_max=int(row.get('Slider Level Max', 5)),
                                image_filename=row.get('Bild', '')
                            )
                            db.session.add(activity)
                        db.session.commit()
                        flash(f'{len(rows)} Aktivität(en) aus CSV importiert!', 'success')

                else:
                    flash('Ungültiges Dateiformat! Nur .xlsx oder .csv erlaubt.', 'danger')
                    return redirect(url_for('import_data'))

                return redirect(url_for('admin'))

            except Exception as e:
                db.session.rollback()
                flash(f'Fehler beim Import: {str(e)}', 'danger')
                return redirect(url_for('import_data'))

        return render_template('admin_import.html')


# ========== HELPER FUNCTIONS ==========

def allowed_file(filename):
    """Prüfe ob Dateiendung erlaubt ist"""
    from app.config import Config
    if '.' not in filename:
        return False
    return filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS


def sanitize_folder_name(name):
    """Erstelle sicheren Ordnernamen aus Destination-Name"""
    import re
    # Ersetze Umlaute
    replacements = {'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'ß': 'ss',
                   'Ä': 'Ae', 'Ö': 'Oe', 'Ü': 'Ue'}
    for old, new in replacements.items():
        name = name.replace(old, new)
    # Entferne alle Zeichen außer Buchstaben, Zahlen, Bindestriche und Unterstriche
    name = re.sub(r'[^\w\s-]', '', name)
    # Ersetze Leerzeichen durch Unterstriche
    name = re.sub(r'[-\s]+', '_', name)
    return name.lower()


def save_uploaded_file(file, subfolder='general'):
    """Speichere hochgeladene Datei"""
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)

        # Erstelle Zielordner falls nicht vorhanden
        upload_path = os.path.join(Config.UPLOAD_FOLDER, subfolder)
        os.makedirs(upload_path, exist_ok=True)

        # Speichere Datei
        file_path = os.path.join(upload_path, filename)
        file.save(file_path)

        # Gib relativen Pfad zurück (für Datenbank)
        return os.path.join(subfolder, filename)

    return None


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True, host='0.0.0.0', port=5000)
