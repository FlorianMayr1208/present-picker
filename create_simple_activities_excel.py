#!/usr/bin/env python3
"""
Create a simple Excel file with activity titles, destinations, and descriptions
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load data
with open('data/activities.json', 'r', encoding='utf-8') as f:
    activities = json.load(f)

with open('data/destinations.json', 'r', encoding='utf-8') as f:
    destinations = json.load(f)

# Create destination lookup
dest_lookup = {d['id']: d['name'] for d in destinations}

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Activities"

# Header styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)

# Set headers
headers = ["Destination", "Kategorie", "Aktivität", "Beschreibung"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data
row_num = 2
for activity in activities:
    destination_name = dest_lookup.get(activity['destination_id'], 'Unknown')
    category_title = activity['title']

    # Add sub-items
    if 'sub_items' in activity and activity['sub_items']:
        for sub_item in activity['sub_items']:
            ws.cell(row=row_num, column=1, value=destination_name)
            ws.cell(row=row_num, column=2, value=category_title)
            ws.cell(row=row_num, column=3, value=sub_item['title'])
            ws.cell(row=row_num, column=4, value=sub_item.get('description', ''))

            # Alignment
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical="top")
            ws.cell(row=row_num, column=2).alignment = Alignment(vertical="top")
            ws.cell(row=row_num, column=3).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_num, column=4).alignment = Alignment(vertical="top", wrap_text=True)

            row_num += 1

# Adjust column widths
ws.column_dimensions['A'].width = 25  # Destination
ws.column_dimensions['B'].width = 30  # Kategorie
ws.column_dimensions['C'].width = 40  # Aktivität
ws.column_dimensions['D'].width = 80  # Beschreibung

# Freeze header row
ws.freeze_panes = 'A2'

# Save
output_file = 'activities_simple.xlsx'
wb.save(output_file)
print(f"✓ Excel file created: {output_file}")
print(f"  Total activities: {row_num - 2}")
