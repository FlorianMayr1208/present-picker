import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Read the JSON file
with open('/Users/flo/Desktop/pp/data/activities.json', 'r', encoding='utf-8') as f:
    activities = json.load(f)

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Activities"

# Define headers
headers = [
    'Activity ID',
    'Destination ID',
    'Category Title',
    'Category Description',
    'Category Image',
    'Category Default Selected',
    'Sub-Item ID',
    'Sub-Item Title',
    'Sub-Item Description',
    'Points',
    'Default Selected',
    'Mandatory',
    'From Parents',
    'Is Spontaneous',
    'Sub-Item Image',
    'Slider Level Min',
    'Slider Level Max'
]

# Style for headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data
row_num = 2
for activity in activities:
    # Get category-level info
    activity_id = activity.get('id', '')
    destination_id = activity.get('destination_id', '')
    category_title = activity.get('title', '')
    category_description = activity.get('description', '')
    category_image = activity.get('image_filename', '')
    category_default_selected = activity.get('default_selected', False)
    slider_level_min = activity.get('slider_level_min', '')
    slider_level_max = activity.get('slider_level_max', '')

    # Process sub_items
    sub_items = activity.get('sub_items', [])

    if not sub_items:
        # If no sub_items, write just the category info
        ws.cell(row=row_num, column=1).value = activity_id
        ws.cell(row=row_num, column=2).value = destination_id
        ws.cell(row=row_num, column=3).value = category_title
        ws.cell(row=row_num, column=4).value = category_description
        ws.cell(row=row_num, column=5).value = category_image
        ws.cell(row=row_num, column=6).value = str(category_default_selected)
        ws.cell(row=row_num, column=16).value = slider_level_min
        ws.cell(row=row_num, column=17).value = slider_level_max
        row_num += 1
    else:
        # Write each sub_item
        for sub_item in sub_items:
            ws.cell(row=row_num, column=1).value = activity_id
            ws.cell(row=row_num, column=2).value = destination_id
            ws.cell(row=row_num, column=3).value = category_title
            ws.cell(row=row_num, column=4).value = category_description
            ws.cell(row=row_num, column=5).value = category_image
            ws.cell(row=row_num, column=6).value = str(category_default_selected)

            ws.cell(row=row_num, column=7).value = sub_item.get('id', '')
            ws.cell(row=row_num, column=8).value = sub_item.get('title', '')
            ws.cell(row=row_num, column=9).value = sub_item.get('description', '')
            ws.cell(row=row_num, column=10).value = sub_item.get('points', '')
            ws.cell(row=row_num, column=11).value = str(sub_item.get('default_selected', ''))
            ws.cell(row=row_num, column=12).value = str(sub_item.get('mandatory', ''))
            ws.cell(row=row_num, column=13).value = str(sub_item.get('from_parents', ''))
            ws.cell(row=row_num, column=14).value = str(sub_item.get('is_spontaneous', ''))
            ws.cell(row=row_num, column=15).value = sub_item.get('image_filename', '')

            # Add slider levels if they exist at sub-item level
            sub_slider_min = sub_item.get('slider_level_min', slider_level_min)
            sub_slider_max = sub_item.get('slider_level_max', slider_level_max)
            ws.cell(row=row_num, column=16).value = sub_slider_min
            ws.cell(row=row_num, column=17).value = sub_slider_max

            row_num += 1

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        try:
            if cell.value:
                # Special handling for description columns (wider)
                if column[0].column in [4, 9]:  # Description columns
                    max_length = 60
                    break
                else:
                    max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    adjusted_width = min(max_length + 2, 80)
    ws.column_dimensions[column_letter].width = adjusted_width

# Set row height for header
ws.row_dimensions[1].height = 30

# Apply text wrapping to all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Save the workbook
output_file = '/Users/flo/Desktop/pp/activities_overview.xlsx'
wb.save(output_file)
print(f"Excel file created successfully: {output_file}")
print(f"Total rows: {row_num - 1} (excluding header)")
